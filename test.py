#! Python 3
# - Copy and Paste Ranges using OpenPyXl library
import xlrd
import openpyxl

# Prepare the spreadsheets to copy from and paste too.

# File to be copied
wb = openpyxl.load_workbook("C:/Users/dmohata/PycharmProjects/HelloWorld/F204211T.xlsx")  # Add file name
#wb = xlrd.open_workbook(r'\\Saturn/Test_Data/xPCM/xPHEMT/PCMdata/Y2020_FGT/FGT200305_GTE09/F200305T.xls')  # Add file name

# File to be pasted into
template = openpyxl.load_workbook("tes.xlsx")  # Add file name
temp_sheet = template["Sheet1"] # Add Sheet name

# Copy range of cells as a nested list
# Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    # Loops through selected Rows
    for i in range(startRow, endRow + 1, 1):
        # Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            norm=sheet.cell(row=i, column=j).value
            if isinstance(norm, (float, int)) is True:
                #print(type(norm))
                norm=abs(norm)*1000
            rowSelected.append(norm)
        # Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected


# Paste range
# Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow, endRow + 1, 1):
        countCol = 0
        for j in range(startCol, endCol + 1, 1):
            sheetReceiving.cell(column=i, row=j).value = copiedData[countRow][countCol]
            r = sheetReceiving.cell(column=i, row=j).value
            sheetReceiving.cell(column=i, row=j).value = r
            countCol += 1
        countRow += 1


def waferId(sheet_no):
    #print("Processing...")
    selectedRange = copyRange(5, 3, 5, 3, sheet)
    pastingRange = pasteRange(1, sheet_no, 1, sheet_no, temp_sheet, selectedRange)
    # You can save the template as another file to create a new file here too.s
    template.save("tes.xlsx")
    print("Range copied and pasted!")

def createData(sheet_no):
    print("Processing...")
    selectedRange = copyRange(20, 23, 38, 23, sheet)  # Change the 4 number values
    pastingRange = pasteRange(2, sheet_no, 20, sheet_no, temp_sheet, selectedRange)  # Change the 4 number values
    # You can save the template as another file to create a new file here too.s
    template.save("tes.xlsx")
    print("Range copied and pasted!")
#print(wb.sheetnames)
i = 1
 # Add Sheet name
while i < 5:
    sheet = wb[f'{i}']
    go = waferId(i)
    go2 = createData(i)
    i = i + 1
    if i == 5:
        break



