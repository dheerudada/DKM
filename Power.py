import os
import openpyxl
import xlrd
import pyexcel as p
from cpy_pst import copyRange, pasteRange
from openpyxl.chart import LineChart,Reference, ScatterChart, Series

original_working_directory = os.getcwd()
mask=input('enter mask: ')

if mask[0:2]==('GT'):
    tech ='FGT'
elif mask[0:2]==('GQ'):
    tech ='FGQ'
elif mask[0:2]==('GF'):
    tech ='FPN'
else:
    tech ='FGN'

#Clear results file
template = openpyxl.load_workbook(f'C:/Users/dmohata/PycharmProjects/HelloWorld/tes.xlsx')
#print(template.get_sheet_names())
std=template['Sheet2']
template.remove(std)
#print(template.get_sheet_names())
temp_sheet = template.create_sheet('Sheet2',0)
#print(template.get_sheet_names())
template.save(f'C:/Users/dmohata/PycharmProjects/HelloWorld/tes.xlsx')
#####clear results file done

# do stuff
path=rf'\\Saturn/TestData/Test_data/xPCM/xPHEMT/PCMData/Y2021_{tech}'
new_directory=os.chdir(path)
print(next(os.walk('.'))[1])
folder=next(os.walk('.'))[1]
folder1=[]
new_directory=os.chdir(path)
folder=next(os.walk('.'))[1]
folder1=[]
for nam in folder:
    print(nam[10:15])
    if nam[10:15]==mask:
        content = os.listdir(nam)
        for con in content:
            try:
                if con[7:9]=='P.':
                    folder1.append(nam)
            except IndexError:
                continue

subfolder=[]
i=0
sheet_no = 2
for file in folder1:
    name=f'{file[0]}{file[3:9]}P'
    #print (name)
    subfolder.append(name)
    #print(subfolder)
    os.chdir(path)
    os.chdir(folder1[i])
    try:
        subpath=f'C:/Users/dmohata/PycharmProjects/HelloWorld/{tech}/{subfolder[i]}.xlsx'
        wb = xlrd.open_workbook(f'{subfolder[i]}.xls')
        p.save_book_as(file_name=f'{subfolder[i]}.xls',
                       dest_file_name=subpath)
        wb = openpyxl.load_workbook(subpath)

        i=i+1
        #print(i)

        sh_names=[]
        for names in wb.sheetnames:
            try:
                if isinstance(int(names), int) is True:
                    sh_names.append(names)
            except ValueError:
                continue
        #print(sh_names)

        for sh in sh_names:
            #print(sh)
            ws = wb[sh]
            waferid=ws.cell(row=3, column=3).value
            maskid=ws.cell(row=4, column=3).value
            temp_sheet.cell(row=sheet_no, column=1).value = f'{name}_{waferid}'
            temp_sheet.cell(row=sheet_no, column=2).value = maskid
            #Vgsq_par = ws.cell(row=11, column=2).value
            #Vgsq_data = ws.cell(row=11, column=9).value
            #temp_sheet.cell(row=1, column=2).value = Vgsq_par
            #temp_sheet.cell(row=sheet_no, column=2).value = Vgsq_data
            m=2
            for row in range(0, 20,1):
                param=ws.cell(row=11+row, column=2).value
                data=ws.cell(row=11+row, column=9).value
                temp_sheet.cell(row=1, column=m+1).value=param
                temp_sheet.cell(row=sheet_no, column=m+1).value = data
                m=m+1
            for row in range(0, 37, 1):
                param = ws.cell(row=39 + row, column=2).value
                d=0
                mean=[]
                for d in range(0, 5, 1):
                    data = ws.cell(row=39 + row, column=5+d*3).value
                    if data is not None:
                        mean.append(data)
                        print(type(data))
                temp_sheet.cell(row=1, column=m + 1).value = param
                try:
                    temp_sheet.cell(row=sheet_no, column=m + 1).value = sum(mean)/len(mean)
                except ZeroDivisionError:
                    pass
                m = m + 1
            try:
                # temp_sheet.cell(column=1, row=sheet_no).value=f'{name}_{waferid}'
                # template.save(f'C:/Users/dmohata/PycharmProjects/HelloWorld/tes.xlsx')
                sheet_no = sheet_no + 1
                print(sheet_no)
            except NameError:
                print('parameter does not exist')
                break
    except FileNotFoundError:
        i=i+1
        continue
##Chart plotting starts here
x= Reference(temp_sheet, min_col=1, min_row=2,
                   max_col=1, max_row=temp_sheet.max_row)
y = Reference(temp_sheet, min_col=3, min_row=1,
                   max_col=temp_sheet.max_column, max_row=temp_sheet.max_row)
# Create object of LineChart class
#print(y)
chart = LineChart()
chart.add_data(y,titles_from_data=True)
chart.set_categories(x)
#a=Series(values=y, title_from_data=True)
#chart.append(a)
# set the title of the chart
chart.title = " LINE-CHART "
# set the title of the x-axis
chart.x_axis.title = " X-AXIS "
# set the title of the y-axis
chart.y_axis.title = "  "
# add chart to the sheet
# the top-left corner of a chart
# is anchored to cell E2 .
temp_sheet.add_chart(chart, "E2")
template.save(f'C:/Users/dmohata/PycharmProjects/HelloWorld/tes.xlsx')

#changeback to original working directory
os.chdir(original_working_directory)
print(next(os.walk('.'))[1])
