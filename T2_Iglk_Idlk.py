import os
import openpyxl
import xlrd
import pyexcel as p
from cpy_pst import copyRange, pasteRange
from openpyxl.chart import LineChart,Reference, ScatterChart, Series
from sub import where, clearsheet


original_working_directory = os.getcwd() ## save current directory
mask=input('enter mask: ') ## get mask info
#tech=where(mask) ## run subroutine to refer correct pcm folder by technology
tech="FGQ"
clearsheet() ## run subroutine to clear results file
template = openpyxl.load_workbook(f'C:/Users/dmohata/PycharmProjects/HelloWorld/tes.xlsx')
temp_sheet = template.create_sheet('Sheet2',0)
# do stuff
path=rf'\\Saturn/TestData/Test_data/xPCM/xPHEMT/PCMData/Y2020_{tech}'
new_directory=os.chdir(path)
#print(next(os.walk('.'))[1])
folder=next(os.walk('.'))[1]
folder1=[]
for nam in folder:
    if nam[10:15] == mask:
        content = os.listdir(nam)
        for con in content:
            try:
                if con[7:9]=='G.':
                    folder1.append(nam)
            except IndexError:
                continue
print(folder1)
subfolder=[]
i=0
sheet_no = 2
for file in folder1:
    name=f'{file[0]}{file[3:9]}G'
    #print (name)
    subfolder.append(name)
    #print(subfolder)
    os.chdir(path)
    os.chdir(folder1[i])
    #print(i)

    subpath=f'C:/Users/dmohata/PycharmProjects/HelloWorld/{tech}/{subfolder[i]}.xlsx'
    wb = xlrd.open_workbook(f'{subfolder[i]}.xls')
    p.save_book_as(file_name=f'{subfolder[i]}.xls',
                   dest_file_name=subpath)
    wb = openpyxl.load_workbook(subpath)

    #i=i+1
    #print(i)

    sh_names=[]
    for names in wb.sheetnames:
        try:
            if isinstance(int(names), int) is True:
                sh_names.append(names)
        except ValueError:
            continue
    #print(sh_names)
    print(sh_names)
    sheets = wb.sheetnames
    ws0=wb[sheets[0]]

    for sh in sh_names:
        #print(sh)
        ws = wb[sh]
        waferid=ws.cell(row=3, column=5).value
        temp_sheet.cell(row=sheet_no, column=1).value = f'{name}_{waferid}'
        epid = ws0.cell(row=8 + int(sh)-1, column=4).value
        date = ws.cell(row=3, column=11).value
        try:
            if epid[0:2]=='MA':
                epid='IQE J+'
            elif epid[0:2]=='NA':
                epid='Cree J+'
        except TypeError:
            epid='IQE J+'
            continue
        temp_sheet.cell(row=sheet_no, column=2).value = epid
        temp_sheet.cell(row=sheet_no, column=3).value = date
        m=1
        for row in range(0, 2, 1):
            param = ws.cell(row=35+row, column=1).value
            #print(param)
            data = ws.cell(row=35+row, column=5).value
            temp_sheet.cell(row=1, column=m+3).value = param
            temp_sheet.cell(row=sheet_no, column=m+3).value = data
            m=m+1
        sheet_no = sheet_no + 1
        print(sheet_no)
    i=i+1
##Chart plotting starts here

# Create object of LineChart class
#print(y)
pos=["E5", "E10"]
k=0
for k in range(0,temp_sheet.max_column-1,1):
    chart= LineChart()
    x = Reference(temp_sheet, min_col=1, min_row=2,
                  max_col=1, max_row=temp_sheet.max_row)
    y = Reference(temp_sheet, min_col=2+k, min_row=1,
                  max_col=2+k, max_row=temp_sheet.max_row)
    chart.add_data(y,titles_from_data=True)
    chart.set_categories(x)
    #a=Series(values=y, title_from_data=True)
    #chart.append(a)
    # set the title of the chart
    chart.title = " LINE-CHART "
    # set the title of the x-axis
    chart.x_axis.title = " X-AXIS "
    # set the title of the y-axis
    chart.y_axis.title = "  "
    # add chart to the sheet
    # the top-left corner of a chart
    # is anchored to cell E2 .
    chart.style = 1
    line=chart.series[0]
    line.marker.symbol = "circle"
    temp_sheet.add_chart(chart, "E5")
template.save(f'C:/Users/dmohata/PycharmProjects/HelloWorld/tes.xlsx')

#changeback to original working directory
os.chdir(original_working_directory)
print(next(os.walk('.'))[1])
