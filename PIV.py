import os
import openpyxl
import xlrd
import pyexcel as p
from cpy_pst import copyRange, pasteRange
from openpyxl.chart import LineChart,Reference, ScatterChart, Series

original_working_directory = os.getcwd()
mask=input('enter mask: ')
if mask[0:2]==('GT'):
    tech ='FGT'
elif mask[0:2]==('GQ'):
    tech ='FGQ'
elif mask[0:2]==('GF'):
    tech ='FPN'
else:
    tech ='FGN'

#Clear results file
template = openpyxl.load_workbook(f'C:/Users/dmohata/PycharmProjects/HelloWorld/tes.xlsx')
#print(template.get_sheet_names())
std=template['Sheet1']
template.remove(std)
#print(template.get_sheet_names())
temp_sheet = template.create_sheet('Sheet1',0)
#print(template.get_sheet_names())
template.save(f'C:/Users/dmohata/PycharmProjects/HelloWorld/tes.xlsx')
#####clear results file done

# do stuff
path=rf'\\Saturn/TestData/Test_data/xPCM/xPHEMT/PCMData/Y2020_{tech}'
new_directory=os.chdir(path)
print(next(os.walk('.'))[1])
folder=next(os.walk('.'))[1]
folder1=[]
for nam in folder:
    if nam[10:15]==mask:
        folder1.append(nam)

subfolder=[]
i=0
sheet_no = 2
for file in folder1:
    name=f'{file[0]}{file[3:9]}PL'
    #print (name)
    subfolder.append(name)
    #print(subfolder)
    os.chdir(path)
    os.chdir(folder1[i])
    try:
        subpath=f'C:/Users/dmohata/PycharmProjects/HelloWorld/{tech}/{subfolder[i]}.xlsx'
        wb = xlrd.open_workbook(f'{subfolder[i]}.xls')
        p.save_book_as(file_name=f'{subfolder[i]}.xls',
                       dest_file_name=subpath)
        wb = openpyxl.load_workbook(subpath)

        i=i+1
        #print(i)
        sh_names=[]
        for names in wb.sheetnames:
            try:
                if isinstance(int(names), int) is True:
                    sh_names.append(names)
            except ValueError:
                continue
        #print(sh_names)

        for sh in sh_names:
            #print(sh)
            ws = wb[sh]
            waferid=ws.cell(row=3, column=5).value
            temp_sheet.cell(row=sheet_no, column=1).value = f'{name}_{waferid}'

            m=1
            for row in range(0, 4,1):
                param=ws.cell(row=17+row*17, column=1).value
                data_ref=ws.cell(row=17+row*17, column=5).value
                #data=(ws.cell(row=14+row*17, column=5).value-data_ref)/data_ref
                temp_sheet.cell(row=1, column=m+1).value=param
                temp_sheet.cell(row=sheet_no, column=m+1).value = data_ref
                m=m+1
            try:
                #temp_sheet.cell(column=1, row=sheet_no).value=f'{name}_{waferid}'
                #template.save(f'C:/Users/dmohata/PycharmProjects/HelloWorld/tes.xlsx')
                sheet_no = sheet_no + 1
                print(sheet_no)
            except NameError:
                print('parameter does not exist')
                break
    except FileNotFoundError:
        i=i+1
        continue

for item in range(0,4):
    for row in range(2,temp_sheet.max_row+1):
        base=temp_sheet.cell(row=row, column=2).value
        #temp_sheet.cell(row=row, column=6+item).value = (temp_sheet.cell(row=row, column=2+item).value-base)/base*100
        temp_sheet.cell(row=row, column=6 + item).value = 2/(temp_sheet.cell(row=row, column=2 + item).value)*1000
temp_sheet.cell(row=1, column=6).value='Ron_Std'
temp_sheet.cell(row=1, column=7).value='Ron_Glag'
temp_sheet.cell(row=1, column=8).value='Ron_Dlag'
temp_sheet.cell(row=1, column=9).value='Ron_Mem'
##Chart plotting starts here
x= Reference(temp_sheet, min_col=1, min_row=2,
                   max_col=1, max_row=temp_sheet.max_row)
y = Reference(temp_sheet, min_col=2, min_row=1,
                   max_col=2, max_row=temp_sheet.max_row)
# Create object of LineChart class
#print(y)
chart = LineChart()
chart.add_data(y,titles_from_data=True)
chart.set_categories(x)
#a=Series(values=y, title_from_data=True)
#chart.append(a)
# set the title of the chart
chart.title = f" LINE-CHART_{mask} "
# set the title of the x-axis
chart.x_axis.title = " Lot/Wafer Id "
# set the title of the y-axis
chart.y_axis.title = " Rdon, Ohm.mm "
# add chart to the sheet
# the top-left corner of a chart
# is anchored to cell E2 .
temp_sheet.add_chart(chart, "E10")
template.save(f'C:/Users/dmohata/PycharmProjects/HelloWorld/tes.xlsx')

#changeback to original working directory
os.chdir(original_working_directory)
print(next(os.walk('.'))[1])
