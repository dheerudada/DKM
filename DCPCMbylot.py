import os
import openpyxl
import xlrd
import pyexcel as p
from cpy_pst import copyRange, pasteRange
from openpyxl.chart import LineChart,Reference

original_working_directory = os.getcwd()
#lotid = input('enter lot id: ')
parameter=input('give parameter name:')
UL=float(input('upper limit: '))
LL=float(input('lower limit: '))
NF=float(input('normalization factor: '))

#Clear results file
template = openpyxl.load_workbook(f'C:/Users/dmohata/PycharmProjects/HelloWorld/tes.xlsx')
#print(template.get_sheet_names())
std=template['Sheet1']
template.remove(std)
#print(template.get_sheet_names())
temp_sheet = template.create_sheet('Sheet1',0)
#print(template.get_sheet_names())
template.save(f'C:/Users/dmohata/PycharmProjects/HelloWorld/tes.xlsx')
#####clear results file done

# do stuff
new_directory=os.chdir(r'\\Saturn/TestData/Test_data/xPCM/xPHEMT/PCMData/Y2020_FGQ')
folder=next(os.walk('.'))[1]
folder1=[]
for nam in folder:
    #print(nam[10:15])
    if nam[0:15]=='FGQ203907_GQE33':
        content = os.listdir(nam)
        for con in content:
            try:
                if con[7:9]=='T.':
                    folder1.append(nam)
            except IndexError:
                continue
print(folder1)
subfolder=[]
i=0
sheet_no = 2
for file in folder1:
    if file==file:
    #if file[3:9]==lotid[3:9]:
    #    print(file[3:9], lotid[3:9])
        name=f'{file[0]}{file[3:9]}T'
        #print (name)
        subfolder.append(name)
        #print(subfolder)
        os.chdir(r'\\Saturn/TestData/Test_data/xPCM/xPHEMT/PCMData/Y2020_FGQ')
        os.chdir(file)
        #os.chdir(folder[i])
        wb = xlrd.open_workbook(f'{subfolder[i]}.xls')
        p.save_book_as(file_name=f'{subfolder[i]}.xls',
                   dest_file_name=f'C:/Users/dmohata/PycharmProjects/HelloWorld/FGQ/{subfolder[i]}.xlsx')
        wb = openpyxl.load_workbook(f'C:/Users/dmohata/PycharmProjects/HelloWorld/FGQ/{subfolder[i]}.xlsx')
        i=i+1
        #print(i)
        
        sh_names=[]
        for names in wb.sheetnames:
            try:
                if isinstance(int(names), int) is True:
                    sh_names.append(names)
            except ValueError:
                continue
        #print(sh_names)
        
        for sh in sh_names:
            #print(sh)
            ws = wb[sh]
            waferid=ws.cell(row=3, column=5).value
            date = ws.cell(row=3, column=11).value
            maskid = ws.cell(row=4, column=11).value
        
            param_rows = []
            for row in range(2, ws.max_row + 1):
                if (ws.cell(row=row, column=20).value == parameter):
                    #print("Processing...")
                    datasize=ws.max_column-20
                    #print(datasize)
                    selectedRange = copyRange(20,row, ws.max_column, row, ws, UL, LL, NF)  # Change the 4 number values
                    pastingRange = pasteRange(5, sheet_no, datasize, sheet_no, temp_sheet,selectedRange)  # Change the 4 number values
                    #print("Range copied and pasted!")
            try:
                temp_sheet.cell(row=1, column=sheet_no).value=ws.cell(row=3, column=2).value
                temp_sheet.cell(row=2, column=sheet_no).value = waferid
                temp_sheet.cell(row=3, column=sheet_no).value = date
                temp_sheet.cell(row=4, column=sheet_no).value = maskid
                #template.save(f'C:/Users/dmohata/PycharmProjects/HelloWorld/tes.xlsx')
                sheet_no = sheet_no + 1
                print(sheet_no)
            except NameError:
                print('parameter does not exist')
                break
temp_sheet.cell(row=1, column=1).value="LotId"
temp_sheet.cell(row=2, column=1).value="WaferId"
temp_sheet.cell(row=3, column=1).value="Date"
temp_sheet.cell(row=4, column=1).value="MaskID"
k=4
datasize=temp_sheet.max_row
print(datasize)
while k<datasize+1:
    temp_sheet.cell(row=k, column=1).value=k
    print(k)
    if k==datasize+1:
        break
    k=k+1

##Chart plotting starts here
x= Reference(temp_sheet, min_col=1, min_row=2,
                   max_col=2, max_row=temp_sheet.max_row)
y = Reference(temp_sheet, min_col=6, min_row=1,
                   max_col=temp_sheet.max_column, max_row=temp_sheet.max_row)
# Create object of LineChart class
#print(y)
chart = LineChart()
chart.add_data(y,titles_from_data=True)
chart.set_categories(x)
#a=Series(values=y, title_from_data=True)
#chart.append(a)
# set the title of the chart
chart.title = " LINE-CHART "
# set the title of the x-axis
chart.x_axis.title = " X-AXIS "
# set the title of the y-axis
chart.y_axis.title = "  "
# add chart to the sheet
# the top-left corner of a chart
# is anchored to cell E2 .
temp_sheet.add_chart(chart, "E2")
template.save(f'C:/Users/dmohata/PycharmProjects/HelloWorld/tes.xlsx')

#changeback to original working directory
os.chdir(original_working_directory)
print(next(os.walk('.'))[1])